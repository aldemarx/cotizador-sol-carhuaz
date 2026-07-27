"""
Lee PLANTILLA_IMPORTACION_SOL_DE_CARHUAZ.xlsx y genera src/data/seed.json.

Corre los mismos controles que la hoja VALIDACION del propio archivo.
Este script es la base del importador real del backend (ver documento
de traspaso, seccion 7.6) -- no es codigo desechable.
"""
import json
import sys
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(r"C:\Users\ALDEMAR\Downloads\PLANTILLA_IMPORTACION_SOL_DE_CARHUAZ.xlsx")
OUT_PATH = Path(__file__).resolve().parent.parent / "src" / "data" / "seed.json"

PARAMETROS_NUMERICOS = {
    "cargo_fijo", "descuento_campania", "descuento_contado", "inicial",
    "cuotas_referencia", "cuotas_min", "cuotas_max", "descuento_max_vendedor",
}

COLUMNAS_LOTES = [
    "CODIGO_LOTE", "MANZANA", "LOTE", "AREA_M2", "CODIGO_UBICACION",
    "PRECIO_M2", "MEDIDA_FRENTE", "ENTREGA_FISICA", "ESTADO",
    "PRECIO_LISTA_OVERRIDE", "OBSERVACION",
]


def leer_parametros(wb):
    ws = wb["PARAMETROS"]
    parametros = {}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        clave, valor = row[0], row[1]
        if not clave:
            continue
        if clave in PARAMETROS_NUMERICOS and valor is not None:
            valor = float(valor) if isinstance(valor, float) or "." in str(valor) else int(valor)
        parametros[clave] = valor
    return parametros


def leer_catalogo_ubicacion(wb):
    ws = wb["CATALOGO_UBICACION"]
    catalogo = set()
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        codigo = row[0]
        if codigo:
            catalogo.add(str(codigo).strip().upper())
    return catalogo


def leer_lotes(wb, catalogo_ubicacion, errores, avisos):
    ws = wb["LOTES"]
    lotes = []
    codigos_vistos = set()

    for r in range(5, ws.max_row + 1):
        valores = [ws.cell(row=r, column=c).value for c in range(1, len(COLUMNAS_LOTES) + 1)]
        if valores[0] is None:
            continue
        fila = dict(zip(COLUMNAS_LOTES, valores))
        codigo = str(fila["CODIGO_LOTE"]).strip()

        if codigo in codigos_vistos:
            errores.append(f"fila {r}: CODIGO_LOTE '{codigo}' duplicado")
        codigos_vistos.add(codigo)

        area = fila["AREA_M2"]
        if not isinstance(area, (int, float)) or area < 50 or area > 600:
            errores.append(f"fila {r} ({codigo}): AREA_M2 invalida o fuera de rango (50-600): {area}")

        ubicacion = str(fila["CODIGO_UBICACION"] or "").strip().upper()
        if ubicacion not in catalogo_ubicacion:
            errores.append(f"fila {r} ({codigo}): CODIGO_UBICACION '{ubicacion}' no esta en el catalogo")

        precio_m2 = fila["PRECIO_M2"]
        if not isinstance(precio_m2, (int, float)):
            errores.append(f"fila {r} ({codigo}): PRECIO_M2 invalido: {precio_m2}")
        elif precio_m2 < 200 or precio_m2 > 300:
            avisos.append(f"fila {r} ({codigo}): PRECIO_M2 fuera de 200-300 (dato real, no error): {precio_m2}")

        entrega = fila["ENTREGA_FISICA"]
        if not isinstance(entrega, int) or entrega < 2025 or entrega > 2030:
            errores.append(f"fila {r} ({codigo}): ENTREGA_FISICA fuera de 2025-2030: {entrega}")

        estado = str(fila["ESTADO"] or "").strip().upper()
        if estado not in {"DISPONIBLE", "VENDIDO", "RESERVADO"}:
            errores.append(f"fila {r} ({codigo}): ESTADO invalido: {estado}")

        lotes.append({
            "codigo_lote": codigo,
            "manzana": str(fila["MANZANA"]).strip(),
            "lote": str(fila["LOTE"]).strip(),
            "area_m2": float(area) if isinstance(area, (int, float)) else None,
            "codigo_ubicacion": ubicacion,
            "precio_m2": float(precio_m2) if isinstance(precio_m2, (int, float)) else None,
            "medida_frente": fila["MEDIDA_FRENTE"],
            "entrega_fisica": entrega,
            "estado": estado,
            "precio_total_override": fila["PRECIO_LISTA_OVERRIDE"],
            "observacion": fila["OBSERVACION"],
        })

    return lotes


def main():
    if not EXCEL_PATH.exists():
        print(f"No se encontro el archivo: {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    parametros = leer_parametros(wb)
    catalogo_ubicacion = leer_catalogo_ubicacion(wb)

    errores = []
    avisos = []
    lotes = leer_lotes(wb, catalogo_ubicacion, errores, avisos)

    print(f"Lotes leidos: {len(lotes)}")
    print(f"Parametros: {parametros}")

    if avisos:
        print(f"\nAvisos ({len(avisos)}, informativos, no bloquean):")
        for a in avisos:
            print(f"  - {a}")

    if errores:
        print(f"\nERRORES ({len(errores)}):", file=sys.stderr)
        for e in errores:
            print(f"  - {e}", file=sys.stderr)
        print("\nNo se genero seed.json porque hay errores.", file=sys.stderr)
        sys.exit(1)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    seed = {"parametros": parametros, "lotes": lotes}
    OUT_PATH.write_text(json.dumps(seed, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n0 errores. seed.json generado en: {OUT_PATH}")


if __name__ == "__main__":
    main()
